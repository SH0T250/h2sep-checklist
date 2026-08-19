#!/usr/bin/env python3
"""Turn the tracked contact-list workbook into data/contacts.json for the app.

The workbook (docs/H2SEP-PhII_Contact-List_*.xlsx) stays the thing a human
edits — this only projects it into the shape contacts.html renders. Run it
after every contact change, commit both files together:

    python3 tools/build_contacts_json.py

Phone and email cells hold more than one value often enough that splitting is
worth doing here rather than in the browser: "(830) 609-9090; C: (956) 376-7054"
becomes two entries, each with its own label, so the page can render one
tap-to-call row per number.
"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "contacts.json"
HDR_ROW = 5
COLS = ["category", "org", "scope", "name", "title", "phone", "email",
        "address", "procore", "notes"]


def newest_workbook() -> Path:
    books = sorted((ROOT / "docs").glob("H2SEP-PhII_Contact-List_*.xlsx"))
    if not books:
        raise SystemExit("no contact-list workbook under docs/")
    return books[-1]


def split_labelled(cell: str) -> list[dict]:
    """"Office: (555) 111-2222; C: (555) 333-4444" -> [{label, value}, ...]"""
    out = []
    for part in (p.strip() for p in str(cell).split(";")):
        if not part:
            continue
        m = re.match(r"^([A-Za-z][A-Za-z ./&-]{0,28}?)\s*:\s*(.+)$", part)
        if m:
            out.append({"label": m.group(1).strip(), "value": m.group(2).strip()})
        else:
            out.append({"label": "", "value": part})
    return out


def main() -> None:
    book = newest_workbook()
    ws = openpyxl.load_workbook(book, data_only=True).active

    contacts = []
    for row in ws.iter_rows(min_row=HDR_ROW + 1, max_row=ws.max_row, values_only=True):
        rec = {k: (str(v).strip() if v is not None else "")
               for k, v in zip(COLS, row)}
        if not any(rec.values()):
            continue
        rec["phones"] = split_labelled(rec.pop("phone"))
        rec["emails"] = split_labelled(rec.pop("email"))
        # An open slot with nothing but a note is a real row on the sheet
        # (the engineer-of-record gap) — keep it so the app shows the gap too.
        contacts.append(rec)

    # Category order mirrors the workbook's dropdown, so the app reads in the
    # same sequence as the spreadsheet rather than alphabetically.
    order, seen = [], set()
    for c in contacts:
        if c["category"] and c["category"] not in seen:
            seen.add(c["category"])
            order.append(c["category"])

    payload = {
        "source": book.name,
        "count": len(contacts),
        "categoryOrder": order,
        "contacts": contacts,
    }
    OUT.write_text(json.dumps(payload, indent=1, ensure_ascii=False) + "\n")
    print(f"wrote {OUT.relative_to(ROOT)}  contacts: {len(contacts)}  from {book.name}")


if __name__ == "__main__":
    main()
