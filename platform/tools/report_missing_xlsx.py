import json, re
from collections import OrderedDict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

import sys; SP = sys.argv[1] if len(sys.argv) > 1 else '.'   # folder holding open-issues.json from report_open_issues.mjs; the workbook lands there too
rows = json.load(open(SP + '/open-issues.json'))
DATE = '2026-09-02'
OUT = SP + '/H2SEP_Missing-FFE-Guest-Rooms_2026-09-02.xlsx'

def roomsort(a):
    m = re.match(r'^(\d+)(.*)$', a); return (int(m.group(1)), m.group(2)) if m else (0, a)
def guest(r): return not str(r['type']).startswith('space-') and r['type'] not in ('mep-punch',)
def tag(r): return r['code'] or ''
def item(r): return r['label']
def tlabel(r): return r['typeLabel'] or r['type']

miss = sorted([r for r in rows if r['issue'].startswith('MISSING') and guest(r)], key=lambda r: (r['floor'], roomsort(r['docId']), r['category'] or '', tag(r), item(r)))
other = sorted([r for r in rows if not r['issue'].startswith('MISSING') and guest(r)], key=lambda r: (r['floor'], roomsort(r['docId']), r['category'] or '', tag(r), item(r)))
floors = sorted(set(r['floor'] for r in miss))

FONT = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='1F3A5F')
HDR_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=10)
BODY = Font(name=FONT, size=10); BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True, color='1F3A5F'); SUB = Font(name=FONT, size=9, italic=True, color='555555')
WRAP = Alignment(wrap_text=True, vertical='top'); TOP = Alignment(vertical='top')
NUM = '#,##0'

wb = Workbook()
def header(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h); cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = Alignment(vertical='center', wrap_text=True)
    if widths:
        for c, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(c)].width = max(ws.column_dimensions[get_column_letter(c)].width or 0, w)
def put(ws, row, values, wrap=(), bold=False):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v); cell.font = BOLD if bold else BODY
        cell.alignment = WRAP if c in wrap else TOP
        if isinstance(v, (int, float)) and not isinstance(v, bool): cell.number_format = NUM
def table(ws, ref, name):
    t = Table(displayName=name, ref=ref); t.tableStyleInfo = TableStyleInfo(name='TableStyleLight1', showRowStripes=True); ws.add_table(t)

# ---------------- Summary ----------------
ws0 = wb.active; ws0.title = 'Summary'
ws0['A1'] = 'H2SEP Missing FF&E Items, Guest Rooms'; ws0['A1'].font = TITLE
ws0['A2'] = f'Home2 Suites by Hilton, Eagle Pass TX. Open MISSING flags on guest-room checklist lines, pulled from the live crew checklist on {DATE}. Common areas and MEP punch lists are not included.'; ws0['A2'].font = SUB
r = 4
header(ws0, r, ['Floor', 'Missing Lines', 'Rooms With Something Missing', 'Total Qty Missing', 'Lines Since Initialed'], [28, 14, 16, 14, 14, 14, 14, 14])
ws0.row_dimensions[r].height = 30
tot = [0, 0, 0, 0]
for fl in floors:
    fr = [x for x in miss if x['floor'] == fl]
    v = [len(fr), len(set(x['docId'] for x in fr)), sum(int(x['qty'] or 1) for x in fr), sum(1 for x in fr if x['checked'])]
    r += 1; put(ws0, r, [f'Floor {fl}'] + v); tot = [a + b for a, b in zip(tot, v)]
r += 1; put(ws0, r, ['All floors'] + tot, bold=True)

r += 3; ws0.cell(row=r - 1, column=1, value='Missing lines by room type and floor').font = BOLD
header(ws0, r, ['Room Type'] + [f'Floor {f}' for f in floors] + ['All Floors'])
types = [t for t, _ in Counter(tlabel(x) for x in miss).most_common()]
for t in types:
    per = [sum(1 for x in miss if tlabel(x) == t and x['floor'] == fl) for fl in floors]
    r += 1; put(ws0, r, [t] + per + [sum(per)])
r += 1; put(ws0, r, ['All types'] + [sum(1 for x in miss if x['floor'] == fl) for fl in floors] + [len(miss)], bold=True)

r += 3; ws0.cell(row=r - 1, column=1, value='Most-missed items, whole building (number of rooms missing the item, and total pieces)').font = BOLD
header(ws0, r, ['Tag', 'Item', 'Category'] + [f'Floor {f}' for f in floors] + ['Rooms, All Floors', 'Total Qty'])
ws0.column_dimensions['B'].width = 46; ws0.column_dimensions['C'].width = 18
items = Counter((tag(x), item(x), x['category'] or '') for x in miss)
for (tg, it, cat), n in items.most_common():
    per = [sum(1 for x in miss if tag(x) == tg and item(x) == it and x['floor'] == fl) for fl in floors]
    q = sum(int(x['qty'] or 1) for x in miss if tag(x) == tg and item(x) == it)
    r += 1; put(ws0, r, [tg, it, cat] + per + [n, q], wrap=(2,))
r += 2
c = ws0.cell(row=r, column=1, value='How to read this: a "missing line" is one item in one room (two robe hooks missing in one room is one line, two pieces). "Lines Since Initialed" are lines someone has checked off while the Missing flag stayed open; those need a call on whether the flag should be cleared. Every number on this tab was totaled from the Missing tab when the file was built; filter or pivot the Missing tab for anything else.')
c.font = SUB; c.alignment = WRAP; ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); ws0.row_dimensions[r].height = 48
ws0.freeze_panes = 'A4'

# ---------------- By Floor & Type ----------------
ws3 = wb.create_sheet('By Floor & Type')
H3 = ['Floor', 'Room Type', 'Category', 'Tag', 'Item', 'Rooms Missing It', 'Total Qty Missing', 'Rooms']
header(ws3, 1, H3, [7, 26, 20, 12, 48, 12, 12, 64])
groups = OrderedDict()
for x in miss:
    g = groups.setdefault((x['floor'], tlabel(x), x['category'] or '', tag(x), item(x)), {'rooms': [], 'qty': 0})
    g['rooms'].append(x['docId']); g['qty'] += int(x['qty'] or 1)
i = 1
for (fl, ty, cat, tg, it), g in sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1], kv[0][2], kv[0][3], kv[0][4])):
    i += 1; rooms = sorted(set(g['rooms']), key=roomsort)
    put(ws3, i, [fl, ty, cat, tg, it, len(rooms), g['qty'], ', '.join(rooms)], wrap=(5, 8))
table(ws3, f'A1:H{i}', 'ByFloorType'); ws3.freeze_panes = 'A2'
ws3.cell(row=i + 2, column=1, value='One row per floor, room type and item, with the rooms listed. Qty is the per-room quantity on the checklist, added up.').font = SUB

# ---------------- Missing (one row per room line) ----------------
ws = wb.create_sheet('Missing')
H = ['Floor', 'Room', 'Room Type', 'Category', 'Tag', 'Item', 'Qty', 'Flag', 'Checked Off', 'Initials', 'Notes']
header(ws, 1, H, [7, 8, 26, 20, 12, 52, 6, 30, 11, 9, 30])
i = 1
for x in miss:
    i += 1; put(ws, i, [x['floor'], x['docId'], tlabel(x), x['category'] or '', tag(x), item(x), int(x['qty'] or 1), x['issue'], 'Yes' if x['checked'] else 'No', x['initials'] if x['checked'] else '', ''], wrap=(6, 11))
table(ws, f'A1:K{i}', 'MissingLines'); ws.freeze_panes = 'A2'
ws.conditional_formatting.add(f'A2:K{i}', FormulaRule(formula=['$I2="Yes"'], fill=PatternFill('solid', fgColor='E2EFDA')))
ws.cell(row=i + 2, column=1, value=f'Source: live H2SEP checklist database, read {DATE}. Guest-room lines whose flag reads MISSING (or MISSING: with the missing piece named) and has not been resolved. Floor 4 includes the field walk of 2026-09-02. Green rows were initialed after the flag was raised.').font = SUB

# ---------------- Other Open Flags ----------------
ws2 = wb.create_sheet('Other Open Flags')
header(ws2, 1, H, [7, 8, 26, 20, 12, 52, 6, 34, 11, 9, 30])
i = 1
for x in other:
    i += 1; put(ws2, i, [x['floor'], x['docId'], tlabel(x), x['category'] or '', tag(x), item(x), int(x['qty'] or 1), x['issue'], 'Yes' if x['checked'] else 'No', x['initials'] if x['checked'] else '', ''], wrap=(6, 8, 11))
table(ws2, f'A1:K{i}', 'OtherFlags'); ws2.freeze_panes = 'A2'
ws2.cell(row=i + 2, column=1, value='Guest-room lines with any other open flag: In box, Need install, and notes typed in the field (many read "installed, missing shade"). Same source and date.').font = SUB

wb.save(OUT)
print('missing', len(miss), 'other', len(other), 'groups', len(groups), 'items', len(items), 'floors', floors)
print('totals', tot)
