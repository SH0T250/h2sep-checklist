import json, re, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

import sys
SP=sys.argv[1] if len(sys.argv)>1 else '.'   # holds missing_rows.json from report_open_issues.mjs; the workbook lands here too
rows=json.load(open(SP+'/missing_rows.json'))
STAMP='2026-09-03'
OUT=SP+'/H2SEP_Missing-Items-Reorder_2026-09-03.xlsx'

def roomsort(a):
    m=re.match(r'^(S?)(\d+)(.*)$', a)
    return (m.group(1)=='S', int(m.group(2)), m.group(3)) if m else (True,0,a)

FONT='Arial'
NAVY=PatternFill('solid', fgColor='1F3A5F')
HDR=Font(name=FONT,bold=True,color='FFFFFF',size=10)
BODY=Font(name=FONT,size=10); BOLD=Font(name=FONT,size=10,bold=True)
TITLE=Font(name=FONT,size=15,bold=True,color='1F3A5F')
SUB=Font(name=FONT,size=9,italic=True,color='555555')
KPIL=Font(name=FONT,size=9,bold=True,color='FFFFFF')
KPIV=Font(name=FONT,size=16,bold=True,color='1F3A5F')
WRAP=Alignment(wrap_text=True,vertical='top'); TOP=Alignment(vertical='top')
CEN=Alignment(horizontal='center',vertical='center')
THIN=Side(style='thin',color='D5DBE0')
BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
PART_FILL=PatternFill('solid', fgColor='FFF3CD')

wb=Workbook(); wb.remove(wb.active)
grand={'lines':0,'pieces':0,'parts':0,'rooms':set()}

for floor in ['1','2','3','4']:
    fr=[r for r in rows if r['floor']==floor]
    ws=wb.create_sheet(f'Floor {floor}')
    # group by item + exact flag text
    g=collections.OrderedDict()
    for r in fr:
        key=(r['kind'], r['cat'], r['tag'], r['item'], r['flag'])
        e=g.setdefault(key,{'rooms':[], 'qty':0})
        e['rooms'].append(r['room']); e['qty']+=r['qty']
    whole=[(k,v) for k,v in g.items() if k[4]=='MISSING']
    parts=[(k,v) for k,v in g.items() if k[4]!='MISSING']
    whole.sort(key=lambda kv:(kv[0][0]!='Guest room', kv[0][1], kv[0][2] or 'zz', kv[0][3]))
    parts.sort(key=lambda kv:(kv[0][0]!='Guest room', kv[0][1], kv[0][2] or 'zz', kv[0][3]))
    tot_pieces=sum(v['qty'] for k,v in whole)
    tot_part=sum(v['qty'] for k,v in parts)
    rooms_aff=sorted({r['room'] for r in fr}, key=roomsort)

    ws['A1']=f'FLOOR {floor} - MISSING ITEMS TO REORDER'; ws['A1'].font=TITLE
    ws['A2']=(f'Home2 Suites by Hilton, Eagle Pass TX (H2SEP, Triun 24030). Pulled from the live checklist on {STAMP}. '
              f'Guest rooms and common areas, FF&E lines only - the MEP punch list is not included. '
              f'One row per item and flag: "MISSING" means the whole item is not there, anything else is a part or a field note and is listed separately below.')
    ws['A2'].font=SUB; ws['A2'].alignment=WRAP
    ws.merge_cells('A2:H2'); ws.row_dimensions[2].height=28

    # KPI strip
    kpis=[('Items to reorder (pieces)',tot_pieces),('Distinct items',len(whole)),
          ('Rooms affected',len(rooms_aff)),('Part / note lines',tot_part)]
    for i,(lab,val) in enumerate(kpis):
        c1=ws.cell(row=4,column=1+i*2,value=lab); c1.font=KPIL; c1.fill=NAVY; c1.alignment=CEN
        ws.merge_cells(start_row=4,start_column=1+i*2,end_row=4,end_column=2+i*2)
        c2=ws.cell(row=5,column=1+i*2,value=val); c2.font=KPIV; c2.alignment=CEN; c2.number_format='#,##0'
        ws.merge_cells(start_row=5,start_column=1+i*2,end_row=5,end_column=2+i*2)
        for col in (1+i*2, 2+i*2):
            ws.cell(row=4,column=col).border=BOX; ws.cell(row=5,column=col).border=BOX
    ws.row_dimensions[4].height=18; ws.row_dimensions[5].height=24

    H=['Kind','Category','Tag','Item','Flag as written','Rooms Missing It','Qty To Reorder','Rooms']
    hr=7
    for c,h in enumerate(H,1):
        cell=ws.cell(row=hr,column=c,value=h); cell.font=HDR; cell.fill=NAVY
        cell.alignment=Alignment(wrap_text=True,vertical='center'); cell.border=BOX
    ws.row_dimensions[hr].height=26
    cur=[hr]
    def emit(items, tag_fill=None):
        for (kind,cat,tag,item,flag),v in items:
            cur[0]+=1; i=cur[0]
            rl=sorted(set(v['rooms']), key=roomsort)
            vals=[kind,cat,tag,item,flag,len(rl),v['qty'],', '.join(rl)]
            for c,val in enumerate(vals,1):
                cell=ws.cell(row=i,column=c,value=val); cell.font=BODY; cell.border=BOX
                cell.alignment=WRAP if c in (4,5,8) else TOP
                if c in (6,7): cell.number_format='#,##0'; cell.alignment=Alignment(horizontal='center',vertical='top')
                if tag_fill: cell.fill=tag_fill
    emit(whole)
    i=cur[0]; last_whole=i
    i+=1
    tc=ws.cell(row=i,column=1,value='TOTAL - WHOLE ITEMS TO REORDER'); tc.font=BOLD
    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=6)
    ws.cell(row=i,column=7,value=tot_pieces).font=BOLD
    ws.cell(row=i,column=7).number_format='#,##0'; ws.cell(row=i,column=7).alignment=Alignment(horizontal='center')
    for c in range(1,9): ws.cell(row=i,column=c).border=BOX
    total_row=i
    if parts:
        i+=2
        ws.cell(row=i,column=1,value='PARTS AND FIELD NOTES - these are pieces of an item, not the whole item, so order to the note').font=BOLD
        ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=8)
        i+=1
        for c,h in enumerate(H,1):
            cell=ws.cell(row=i,column=c,value=h); cell.font=HDR; cell.fill=NAVY
            cell.alignment=Alignment(wrap_text=True,vertical='center'); cell.border=BOX
        cur[0]=i
        emit(parts, PART_FILL)
        i=cur[0]+1
        ws.cell(row=i,column=1,value='TOTAL - PART / NOTE LINES').font=BOLD
        ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=6)
        ws.cell(row=i,column=7,value=tot_part).font=BOLD
        ws.cell(row=i,column=7).number_format='#,##0'; ws.cell(row=i,column=7).alignment=Alignment(horizontal='center')
        for c in range(1,9): ws.cell(row=i,column=c).border=BOX
    if last_whole>hr:
        t=Table(displayName=f'Reorder{floor}', ref=f'A{hr}:H{last_whole}')
        t.tableStyleInfo=TableStyleInfo(name='TableStyleLight1', showRowStripes=True); ws.add_table(t)
    for c,w in zip(range(1,9),[13,20,11,46,30,10,10,58]): ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes=f'A{hr+1}'
    i+=2
    n=ws.cell(row=i,column=1,value=('How to read this: "Qty To Reorder" adds up the per-room quantity, so a robe hook that takes two per room counts two in each room it is missing from. '
        '"Rooms Missing It" is how many rooms carry that shortage. The Rooms column lists them so a runner can walk straight to them. '
        'Numbers are a snapshot of the live checklist at the time above; re-pull any time and they update.'))
    n.font=SUB; n.alignment=WRAP; ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=8); ws.row_dimensions[i].height=30
    grand['lines']+=len(whole); grand['pieces']+=tot_pieces; grand['parts']+=tot_part
    grand['rooms']|=set(rooms_aff)
    print(f"Floor {floor}: {len(whole)} distinct items, {tot_pieces} pieces to reorder, {tot_part} part/note lines, {len(rooms_aff)} rooms")


# ---------------- Summary: every floor on one tab ----------------
# Austin asked for the four floors combined. One row per item and flag, quantity
# broken out by floor and totalled, biggest shortage first so the long poles for
# the order sit at the top. It goes in front of the floor tabs.
ws=wb.create_sheet('Summary', 0)
g=collections.OrderedDict()
for r in rows:
    key=(r['kind'], r['cat'], r['tag'], r['item'], r['flag'])
    e=g.setdefault(key, {'f':{'1':0,'2':0,'3':0,'4':0}, 'rooms':set(), 'qty':0})
    e['f'][r['floor']]+=r['qty']; e['rooms'].add(r['room']); e['qty']+=r['qty']
whole=[(k,v) for k,v in g.items() if k[4]=='MISSING']
parts=[(k,v) for k,v in g.items() if k[4]!='MISSING']
whole.sort(key=lambda kv:(-kv[1]['qty'], kv[0][1], kv[0][2] or 'zz'))
parts.sort(key=lambda kv:(-kv[1]['qty'], kv[0][1], kv[0][2] or 'zz'))
tot=sum(v['qty'] for k,v in whole); totp=sum(v['qty'] for k,v in parts)
allrooms=set()
for k,v in g.items(): allrooms |= v['rooms']

ws['A1']='H2SEP - MISSING ITEMS TO REORDER - ALL FLOORS'; ws['A1'].font=TITLE
ws['A2']=('Home2 Suites by Hilton, Eagle Pass TX (H2SEP, Triun 24030). Every floor combined, pulled from the live checklist on '
          + STAMP + '. Guest rooms and common areas, FF&E lines only, MEP punch not included. Biggest shortage first. '
          'The four floor tabs carry the same lines with the room numbers listed.')
ws['A2'].font=SUB; ws['A2'].alignment=WRAP
ws.merge_cells('A2:K2'); ws.row_dimensions[2].height=28
for i,(lab,val) in enumerate([('Pieces to reorder',tot),('Distinct items',len(whole)),('Rooms affected',len(allrooms)),('Part / note lines',totp)]):
    c1=ws.cell(row=4,column=1+i*2,value=lab); c1.font=KPIL; c1.fill=NAVY; c1.alignment=CEN
    ws.merge_cells(start_row=4,start_column=1+i*2,end_row=4,end_column=2+i*2)
    c2=ws.cell(row=5,column=1+i*2,value=val); c2.font=KPIV; c2.alignment=CEN; c2.number_format='#,##0'
    ws.merge_cells(start_row=5,start_column=1+i*2,end_row=5,end_column=2+i*2)
    for col in (1+i*2, 2+i*2):
        ws.cell(row=4,column=col).border=BOX; ws.cell(row=5,column=col).border=BOX
ws.row_dimensions[4].height=18; ws.row_dimensions[5].height=24

HS=['Kind','Category','Tag','Item','Flag as written','Floor 1','Floor 2','Floor 3','Floor 4','Rooms Missing It','Total Qty To Reorder']
hr=7
for c,h in enumerate(HS,1):
    cell=ws.cell(row=hr,column=c,value=h); cell.font=HDR; cell.fill=NAVY
    cell.alignment=Alignment(wrap_text=True,vertical='center'); cell.border=BOX
ws.row_dimensions[hr].height=30
row_i=[hr]
def emit_sum(items, fill=None):
    for (kind,cat,tag,item,flag),v in items:
        row_i[0]+=1; i=row_i[0]
        vals=[kind,cat,tag,item,flag,v['f']['1'],v['f']['2'],v['f']['3'],v['f']['4'],len(v['rooms']),v['qty']]
        for c,val in enumerate(vals,1):
            cell=ws.cell(row=i,column=c)
            if c in (6,7,8,9) and val==0: val=None
            cell.value=val; cell.font=BODY; cell.border=BOX
            if c in (6,7,8,9,10,11):
                cell.number_format='#,##0'; cell.alignment=Alignment(horizontal='center',vertical='top')
                if c==11: cell.font=BOLD
            else:
                cell.alignment=WRAP if c in (4,5) else TOP
            if fill: cell.fill=fill
emit_sum(whole)
last=row_i[0]
row_i[0]+=1; i=row_i[0]
tc=ws.cell(row=i,column=1,value='TOTAL - WHOLE ITEMS TO REORDER, ALL FLOORS'); tc.font=BOLD
ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=5)
fl_tot=[sum(v['f'][f] for k,v in whole) for f in ('1','2','3','4')]
rooms_w=set()
for k,v in whole: rooms_w |= v['rooms']
for c,val in zip(range(6,12), fl_tot+[len(rooms_w), tot]):
    cell=ws.cell(row=i,column=c,value=val); cell.font=BOLD; cell.number_format='#,##0'
    cell.alignment=Alignment(horizontal='center')
for c in range(1,12): ws.cell(row=i,column=c).border=BOX
if parts:
    row_i[0]+=2; i=row_i[0]
    ws.cell(row=i,column=1,value='PARTS AND FIELD NOTES - a piece of an item, not the whole item, so order to the note').font=BOLD
    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=11)
    row_i[0]+=1; i=row_i[0]
    for c,h in enumerate(HS,1):
        cell=ws.cell(row=i,column=c,value=h); cell.font=HDR; cell.fill=NAVY
        cell.alignment=Alignment(wrap_text=True,vertical='center'); cell.border=BOX
    emit_sum(parts, PART_FILL)
    row_i[0]+=1; i=row_i[0]
    ws.cell(row=i,column=1,value='TOTAL - PART / NOTE LINES, ALL FLOORS').font=BOLD
    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=10)
    cell=ws.cell(row=i,column=11,value=totp); cell.font=BOLD; cell.number_format='#,##0'
    cell.alignment=Alignment(horizontal='center')
    for c in range(1,12): ws.cell(row=i,column=c).border=BOX
if last>hr:
    t=Table(displayName='ReorderAll', ref='A%d:K%d' % (hr,last))
    t.tableStyleInfo=TableStyleInfo(name='TableStyleLight1', showRowStripes=True); ws.add_table(t)
for c,w in zip(range(1,12),[13,20,11,46,30,9,9,9,9,11,13]): ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes='A%d' % (hr+1)
row_i[0]+=2; i=row_i[0]
n=ws.cell(row=i,column=1,value=('How to read this: the four floor columns are quantities, so a blank means that floor is not short of that item. '
  '"Total Qty To Reorder" adds the per-room quantity across the building, which is the number to put on the order. '
  'Open a floor tab for the room numbers behind any line. Snapshot of the live checklist at the time above.'))
n.font=SUB; n.alignment=WRAP; ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=11); ws.row_dimensions[i].height=30
print('Summary: %d item rows, %d pieces, %d part lines, %d rooms' % (len(whole), tot, totp, len(allrooms)))

wb.save(OUT)
print(f"\nBUILDING TOTAL: {grand['pieces']} pieces across {grand['lines']} item rows, {grand['parts']} part lines, {len(grand['rooms'])} rooms")
print('saved', OUT)
