# Austin: "remove the hard code from the columns, I want to be able to change the
# values." Every figure that is arithmetic on other cells becomes a live formula,
# so editing a quantity updates the totals. The figures he types stay as numbers
# and are coloured blue, the standard for an input cell.
from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil, sys
SRC=sys.argv[2] if len(sys.argv)>2 else ''   # the workbook to convert
assert SRC, 'usage: xlsx_unhardcode.py <out dir> <workbook.xlsx>'
SP=sys.argv[1]
OUT=SP+'/H2SEP_FFE_Missing_Items_live.xlsx'
shutil.copy(SRC, OUT)
wb=load_workbook(OUT)
BLUE=Font(name='Arial', size=10, color='0000FF')
changed={'missing_total':0,'ww_count':0,'kpi':0}

# ---- Missing FF&E Items to Order: J = the four floor columns added up ----
ws=wb['Missing FF&E Items to Order']
HDR=4; FIRST=5
LAST=HDR
r=FIRST
while ws.cell(row=r,column=1).value and not str(ws.cell(row=r,column=1).value).startswith('TOTAL'):
    LAST=r; r+=1
TOTROW=r
for r in range(FIRST, LAST+1):
    ws.cell(row=r,column=10).value=f'=SUM(F{r}:I{r})'
    changed['missing_total']+=1
    for c in range(6,10):
        cell=ws.cell(row=r,column=c)
        f=cell.font
        cell.font=Font(name=f.name or 'Arial', size=f.size or 10, bold=f.bold, color='0000FF')
# total row: keep the table-wide sums, make them plain ranges so they survive any row insert
for c,col in zip(range(6,10),'FGHI'):
    ws.cell(row=TOTROW,column=c).value=f'=SUM({col}{FIRST}:{col}{LAST})'
ws.cell(row=TOTROW,column=10).value=f'=SUM(J{FIRST}:J{LAST})'
ws.cell(row=2,column=1).value=(str(ws.cell(row=2,column=1).value or '')+
  '  Blue figures are yours to edit - change a floor quantity and the Total column and the total row follow. Black figures work themselves out.')

# ---- Working Wall Pieces: pieces-per-room count and the four tiles ----
ws=wb['Working Wall Pieces']
HDR=7; FIRST=8
LAST=HDR
r=FIRST
while ws.cell(row=r,column=2).value and not str(ws.cell(row=r,column=1).value or '').startswith('TOTAL'):
    LAST=r; r+=1
TOTROW=r
for r in range(FIRST, LAST+1):
    ws.cell(row=r,column=7).value=f'=COUNTIFS($B${FIRST}:$B${LAST},$B{r},$D${FIRST}:$D${LAST},$D{r})'
    changed['ww_count']+=1
    for c in (5,6):
        cell=ws.cell(row=r,column=c); f=cell.font
        cell.font=Font(name=f.name or 'Arial', size=f.size or 10, bold=f.bold, color='0000FF')
ws.cell(row=TOTROW,column=5).value=f'=COUNTA($E${FIRST}:$E${LAST})'
ws['A5']=f'=COUNTA($E${FIRST}:$E${LAST})'
ws['C5']=f'=SUMPRODUCT(($B${FIRST}:$B${LAST}<>"")/COUNTIF($B${FIRST}:$B${LAST},$B${FIRST}:$B${LAST}&""))'
ws['E5']=f'=SUMPRODUCT(($B${FIRST}:$B${LAST}<>"")/COUNTIFS($B${FIRST}:$B${LAST},$B${FIRST}:$B${LAST}&"",$D${FIRST}:$D${LAST},$D${FIRST}:$D${LAST}&""))'
ws['G5']=f'=COUNTIF($F${FIRST}:$F${LAST},"Opened")'
changed['kpi']=4
ws.cell(row=2,column=1).value=(str(ws.cell(row=2,column=1).value or '')+
  '  Blue figures are yours to edit - add or delete a piece row and the four tiles at the top, the pieces-per-room column and the total all follow.')

# ---- Stored "Extra" FF&E: quantities stay typed, everything else already lives ----
ws=wb['Stored "Extra" FF&E']
r=2
while ws.cell(row=r,column=3).value:
    cell=ws.cell(row=r,column=7); f=cell.font
    cell.font=Font(name=f.name or 'Arial', size=f.size or 10, bold=f.bold, color='0000FF')
    r+=1
LASTS=r-1
for rr in range(2, LASTS+1):
    ws.cell(row=rr,column=8).value=f'=SUMIF($C$2:$C${LASTS},$C{rr},$G$2:$G${LASTS})'
ws.cell(row=LASTS+1,column=7).value=f'=SUM(G2:G{LASTS})'
ws.cell(row=LASTS+1,column=2).value=f'=COUNTA($C$2:$C${LASTS})&" lines across floors 1 to 4"'

wb.save(OUT)
print('Missing tab: %d Total cells now formulas, floor columns blue, total row r%d' % (changed['missing_total'], TOTROW))
print('Working Wall: %d pieces-per-room cells now formulas, 4 tiles now formulas' % changed['ww_count'])
print('Stored Extra: %d rows re-pointed at the live range, quantity column blue' % (LASTS-1))
print('saved', OUT)
