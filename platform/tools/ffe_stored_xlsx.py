"""Austin, 2026-09-03: he walked every floor and wrote the leftover FF&E boxes on
five sheets of notebook paper, then shot them as ALL_Floors_FFE_Stored.pdf. This
turns the transcription in platform/data/ffe-stored-2026-09-03.json into the
workbook he asked for: line-by-line as written, per-code totals, a working wall
breakout, a floor summary, and the flag list.

Two traps the source PDF sets, already handled in the JSON, not here:
  - pages 7 and 5 are ONE sheet (1st floor on top, Room 217 below the "217" heading)
  - page 4 is a re-shoot of page 3 and is NOT counted
Usage: python3 platform/tools/ffe_stored_xlsx.py [outdir]
"""
import json, collections, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, '..', 'data', 'ffe-stored-2026-09-03.json')
OUTDIR = sys.argv[1] if len(sys.argv) > 1 else '.'
STAMP = '2026-09-03'
OUT = os.path.join(OUTDIR, f'H2SEP_FFE-Stored-By-Floor_{STAMP}.xlsx')

doc = json.load(open(SRC))
items, meta = doc['items'], doc['meta']
WW = set(meta['workingWallBaseCodes'])

FONT = 'Arial'
NAVY = PatternFill('solid', fgColor='1F3A5F')
STEEL = PatternFill('solid', fgColor='E8EDF2')
FLAGF = PatternFill('solid', fgColor='FFF3CD')
HDR = Font(name=FONT, bold=True, color='FFFFFF', size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=15, bold=True, color='1F3A5F')
SUB = Font(name=FONT, size=9, italic=True, color='555555')
KPIL = Font(name=FONT, size=9, bold=True, color='FFFFFF')
KPIV = Font(name=FONT, size=16, bold=True, color='1F3A5F')
WRAP = Alignment(wrap_text=True, vertical='top')
TOP = Alignment(vertical='top')
CEN = Alignment(horizontal='center', vertical='center')
THIN = Side(style='thin', color='D5DBE0')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
wb.remove(wb.active)


def head(ws, title, sub, ncols):
    ws['A1'] = title
    ws['A1'].font = TITLE
    ws['A2'] = sub
    ws['A2'].font = SUB
    ws['A2'].alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.row_dimensions[2].height = 30


def kpis(ws, pairs, row, ncols):
    for n, (lab, val) in enumerate(pairs):
        c = 1 + n * 2
        ws.cell(row=row, column=c, value=lab).font = KPIL
        ws.cell(row=row, column=c).fill = NAVY
        ws.cell(row=row, column=c).alignment = CEN
        ws.merge_cells(start_row=row, start_column=c, end_row=row, end_column=min(c + 1, ncols))
        ws.cell(row=row + 1, column=c, value=val).font = KPIV
        ws.cell(row=row + 1, column=c).alignment = CEN
        ws.merge_cells(start_row=row + 1, start_column=c, end_row=row + 1, end_column=min(c + 1, ncols))
    ws.row_dimensions[row + 1].height = 24


def table(ws, cols, rows, start, flagcol=None):
    for j, (name, _) in enumerate(cols, 1):
        cell = ws.cell(row=start, column=j, value=name)
        cell.font = HDR
        cell.fill = NAVY
        cell.alignment = CEN
        cell.border = BOX
    for i, r in enumerate(rows):
        for j, (_, key) in enumerate(cols, 1):
            cell = ws.cell(row=start + 1 + i, column=j, value=r.get(key) if isinstance(r, dict) else r[j - 1])
            cell.font = BODY
            cell.alignment = WRAP if key in ('flag', 'label', 'writtenAs', 'positions') else TOP
            cell.border = BOX
            if flagcol and r.get('flag'):
                cell.fill = FLAGF
    for j, (name, key) in enumerate(cols, 1):
        w = max(len(str(name)), *(len(str((r.get(key) if isinstance(r, dict) else r[j - 1]) or '')) for r in rows)) if rows else len(name)
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 9), 60)
    ws.freeze_panes = ws.cell(row=start + 1, column=1)


# ---- Tab 1: line by line, exactly as written
ws = wb.create_sheet('Line by line')
head(ws, 'EXTRA FF&E IN STORAGE - LINE BY LINE',
     f"{meta['project']}. Walked and written by {meta['capturedBy']} on {meta['capturedOn']}; transcribed from {meta['source']}. "
     f"{meta['tallyConvention']} Highlighted rows carry a flag that needs Austin's confirmation before this feeds the crew app.", 11)
kpis(ws, [('Lines', len(items)), ('Pieces', sum(i['qty'] for i in items)),
          ('Working walls', sum(i['qty'] for i in items if i['isWorkingWall'])),
          ('Flagged', sum(1 for i in items if i['flag']))], 4, 11)
cols = [('ID', 'id'), ('Floor', 'floor'), ('Location', 'location'), ('Item code', 'fullCode'),
        ('Base code', 'code'), ('Position', 'positionCode'), ('Catalog description', 'label'),
        ('Category', 'category'), ('ADA', 'adaTxt'), ('Qty', 'qty'), ('As written on the sheet', 'writtenAs'),
        ('Tally / mark', 'qtyMarks'), ('Source pg', 'sourcePage'), ('Flag', 'flag')]
rows = [dict(r, adaTxt='ADA' if r['ada'] else '') for r in items]
table(ws, cols, rows, 7, flagcol=True)

# ---- Tab 2: per item totals
ws = wb.create_sheet('Item totals')
byc = collections.defaultdict(lambda: {'qty': 0, 'lines': 0, 'floors': set(), 'label': '', 'cat': ''})
for i in items:
    k = i['code'] or i['writtenAs']
    e = byc[k]
    e['qty'] += i['qty']
    e['lines'] += 1
    e['floors'].add(i['floor'])
    e['label'] = e['label'] or i['label'] or i['writtenAs']
    e['cat'] = e['cat'] or i['category']
rows = [{'code': k, 'label': v['label'], 'cat': v['cat'], 'qty': v['qty'], 'lines': v['lines'],
         'floors': ', '.join(str(f) for f in sorted(v['floors']))}
        for k, v in sorted(byc.items(), key=lambda x: (-x[1]['qty'], x[0]))]
head(ws, 'TOTAL QUANTITY BY ITEM', 'Every distinct item across all four floors, highest count first. '
     'Codes with no catalog match are listed under the words Austin wrote.', 6)
kpis(ws, [('Distinct items', len(rows)), ('Total pieces', sum(r['qty'] for r in rows))], 4, 6)
table(ws, [('Item code', 'code'), ('Catalog description', 'label'), ('Category', 'cat'),
           ('Total qty', 'qty'), ('Lines', 'lines'), ('Floors', 'floors')], rows, 7)

# ---- Tab 3: working walls
ws = wb.create_sheet('Working walls')
wwi = [i for i in items if i['isWorkingWall']]
byb = collections.defaultdict(lambda: collections.Counter())
for i in wwi:
    byb[i['code']][i['positionCode'] or '(no position)'] += i['qty']
rows = []
for b in sorted(byb):
    q = sum(byb[b].values())
    a = sum(i['qty'] for i in wwi if i['code'] == b and i['ada'])
    lab = next((i['label'] for i in wwi if i['code'] == b), '')
    rows.append({'code': b, 'label': lab, 'qty': q, 'ada': a, 'std': q - a,
                 'positions': ', '.join(f'{p} x{c}' if c > 1 else p for p, c in sorted(byb[b].items())),
                 'floors': ', '.join(str(f) for f in sorted({i['floor'] for i in wwi if i['code'] == b}))})
head(ws, 'WORKING WALL BREAKOUT',
     'Working wall base codes per the project FF&E catalog: ' + ', '.join(sorted(WW)) + '. '
     'Position codes are room-specific and not interchangeable, so the full position list is kept. '
     'Every GR-316 is accessible by definition (King One Bedroom Suite Accessible).', 7)
kpis(ws, [('Working walls', sum(r['qty'] for r in rows)), ('ADA', sum(r['ada'] for r in rows)),
          ('Standard', sum(r['std'] for r in rows)), ('Base codes', len(rows))], 4, 7)
table(ws, [('Base code', 'code'), ('Catalog description', 'label'), ('Total qty', 'qty'),
           ('ADA', 'ada'), ('Standard', 'std'), ('Positions found', 'positions'), ('Floors', 'floors')], rows, 7)

start = 7 + len(rows) + 3
ws.cell(row=start - 1, column=1, value='EVERY WORKING WALL LINE, AS WRITTEN').font = BOLD
wwrows = [dict(r, adaTxt='ADA' if r['ada'] else '') for r in wwi]
table(ws, [('ID', 'id'), ('Floor', 'floor'), ('Location', 'location'), ('Item code', 'fullCode'),
           ('ADA', 'adaTxt'), ('Qty', 'qty'), ('As written', 'writtenAs'), ('Flag', 'flag')], wwrows, start)

# ---- Tab 4: floor summary
ws = wb.create_sheet('Floor summary')
rows = []
for f in sorted({i['floor'] for i in items}):
    fi = [i for i in items if i['floor'] == f]
    locs = collections.Counter()
    for i in fi:
        locs[i['location']] += i['qty']
    rows.append({'floor': f, 'lines': len(fi), 'qty': sum(i['qty'] for i in fi),
                 'ww': sum(i['qty'] for i in fi if i['isWorkingWall']),
                 'ada': sum(i['qty'] for i in fi if i['ada']),
                 'flags': sum(1 for i in fi if i['flag']),
                 'where': '; '.join(f'{k} ({v})' for k, v in locs.most_common())})
rows.append({'floor': 'TOTAL', 'lines': sum(r['lines'] for r in rows), 'qty': sum(r['qty'] for r in rows),
             'ww': sum(r['ww'] for r in rows), 'ada': sum(r['ada'] for r in rows),
             'flags': sum(r['flags'] for r in rows), 'where': ''})
head(ws, 'STORED FF&E BY FLOOR', 'Where the extra boxes are sitting, floor by floor, with the piece count at each location.', 7)
table(ws, [('Floor', 'floor'), ('Lines', 'lines'), ('Pieces', 'qty'), ('Working walls', 'ww'),
           ('ADA pieces', 'ada'), ('Flags', 'flags'), ('Locations (pieces)', 'where')], rows, 5)

# ---- Tab 5: flags and source notes
ws = wb.create_sheet('Flags and notes')
rows = [{'id': i['id'], 'floor': i['floor'], 'location': i['location'], 'written': i['writtenAs'],
         'qty': i['qty'], 'flag': i['flag'], 'pg': i['sourcePage']} for i in items if i['flag']]
head(ws, 'FLAGS - CONFIRM BEFORE THIS FEEDS THE CREW APP',
     'Every line where the handwriting, the code, or the tally is open to more than one reading. '
     'Nothing here was guessed silently.', 7)
table(ws, [('ID', 'id'), ('Floor', 'floor'), ('Location', 'location'), ('As written', 'written'),
           ('Qty used', 'qty'), ('What needs confirming', 'flag'), ('Source pg', 'pg')], rows, 5)

start = 5 + len(rows) + 3
ws.cell(row=start - 1, column=1, value='STRUCK-THROUGH LINES - NOT COUNTED').font = BOLD
table(ws, [('Floor', 'floor'), ('Location', 'location'), ('As written', 'writtenAs'),
           ('Why excluded', 'reason'), ('Source pg', 'sourcePage')], doc['struckThroughEntries'], start)

start2 = start + len(doc['struckThroughEntries']) + 4
ws.cell(row=start2 - 1, column=1, value='HOW THE PHOTOS MAP TO THE SHEETS').font = BOLD
table(ws, [('Photo', 'k'), ('What it is', 'v')],
      [{'k': k, 'v': v} for k, v in meta['sheetMap'].items()], start2)

wb.save(OUT)
print('wrote', OUT)
print(f"{len(items)} lines, {sum(i['qty'] for i in items)} pieces, "
      f"{sum(i['qty'] for i in items if i['isWorkingWall'])} working walls, "
      f"{sum(1 for i in items if i['flag'])} flagged")
