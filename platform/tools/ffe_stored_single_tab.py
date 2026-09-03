"""Austin, 2026-09-03: he wants the stored-FF&E count as ONE flat sheet he can
filter, with a running total per item across every floor. GR-502 sits on floors
2, 3 and 4, and he wants each of those rows to also show the complete total (4).

The total is a live SUMIF, not a number typed in, so it re-totals if he edits a
qty or adds a row. It keys on the "Item code (base)" column: the catalog base
code where there is one, and the item name where there is none, so "Dishwasher",
"Dishwasher ADA" and "Dishwasher stainless steel" stay three separate totals
instead of collapsing on a blank code.

Row order is the walk order from the sheets, so it still reads against the paper.
Usage: python3 platform/tools/ffe_stored_single_tab.py [outdir]
"""
import json, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, '..', 'data', 'ffe-stored-2026-09-03.json')
OUTDIR = sys.argv[1] if len(sys.argv) > 1 else '.'
STAMP = '2026-09-03'
OUT = os.path.join(OUTDIR, f'H2SEP_FFE-Stored-Totals_{STAMP}.xlsx')

doc = json.load(open(SRC))
items = doc['items']

FONT = 'Arial'
NAVY = PatternFill('solid', fgColor='1F3A5F')
TOTALF = PatternFill('solid', fgColor='E8EDF2')
FLAGF = PatternFill('solid', fgColor='FFF3CD')
HDR = Font(name=FONT, bold=True, color='FFFFFF', size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
NAVYTXT = Font(name=FONT, size=10, bold=True, color='1F3A5F')
WRAP = Alignment(wrap_text=True, vertical='top')
TOP = Alignment(vertical='top')
CEN = Alignment(horizontal='center', vertical='center')
THIN = Side(style='thin', color='D5DBE0')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ('ID', 'id', 7), ('Floor', 'floor', 7), ('Location', 'location', 18),
    ('Item code', 'fullCode', 16), ('Item code (base)', 'itemKey', 17),
    ('Position', 'positionCode', 9), ('Catalog description', 'label', 42),
    ('Category', 'category', 18), ('ADA', 'adaTxt', 6), ('Qty', 'qty', 6),
    ('TOTAL QTY - ALL FLOORS', 'total', 14), ('As written on the sheet', 'writtenAs', 34),
    ('Tally / mark', 'qtyMarks', 11), ('Source pg', 'sourcePage', 9),
    ('Flag', 'flag', 40), ('Confirmed by Austin', 'resolution', 40),
]
KEYCOL = get_column_letter(1 + [c[1] for c in COLS].index('itemKey'))
QTYCOL = get_column_letter(1 + [c[1] for c in COLS].index('qty'))
TOTCOL = get_column_letter(1 + [c[1] for c in COLS].index('total'))

rows = [dict(r,
             itemKey=r['code'] or r['writtenAs'],
             adaTxt='ADA' if r['ada'] else '')
        for r in items]

wb = Workbook()
ws = wb.active
ws.title = 'Stored FF&E'

for j, (name, _, w) in enumerate(COLS, 1):
    c = ws.cell(row=1, column=j, value=name)
    c.font = HDR
    c.fill = NAVY
    c.alignment = CEN
    c.border = BOX
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[1].height = 30

first, last = 2, 1 + len(rows)
for i, r in enumerate(rows):
    rn = first + i
    for j, (_, key, _) in enumerate(COLS, 1):
        if key == 'total':
            # live total for this item across every floor, so an edited qty re-totals
            v = f'=SUMIF(${KEYCOL}${first}:${KEYCOL}${last},${KEYCOL}{rn},${QTYCOL}${first}:${QTYCOL}${last})'
        else:
            v = r.get(key)
        c = ws.cell(row=rn, column=j, value=v)
        c.font = BODY
        c.alignment = WRAP if key in ('label', 'writtenAs', 'flag', 'resolution', 'location') else TOP
        c.border = BOX
        if key == 'total':
            c.font = NAVYTXT
            c.alignment = CEN
            c.fill = TOTALF
        if r['flag'] and key in ('id', 'flag'):
            c.fill = FLAGF

tr = last + 1
ws.cell(row=tr, column=1, value='TOTAL').font = BOLD
ws.cell(row=tr, column=3, value=f'{len(rows)} lines across floors 1 to 4').font = BODY
tc = ws.cell(row=tr, column=1 + [c[1] for c in COLS].index('qty'), value=f'=SUM({QTYCOL}{first}:{QTYCOL}{last})')
tc.font = BOLD
tc.alignment = CEN
for j in range(1, len(COLS) + 1):
    ws.cell(row=tr, column=j).fill = TOTALF
    ws.cell(row=tr, column=j).border = BOX

note = tr + 2
ws.cell(row=note, column=1, value=(
    'TOTAL QTY - ALL FLOORS is a live SUMIF on the "Item code (base)" column, so every row of the same item shows the same '
    'complete total and the sheet re-totals if a qty is edited. Items with no catalog code (Dishwasher, Microwave, TV mounts) '
    'total on their written name, so the ADA and stainless variants stay separate. '
    'Source: Austin\'s handwritten floor sheets, 2026-09-03. Highlighted IDs carry a flag.')).font = Font(name=FONT, size=9, italic=True, color='555555')
ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=len(COLS))
ws.row_dimensions[note].height = 28

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}{last}'
wb.save(OUT)
print('wrote', OUT)
print(f'{len(rows)} rows, {sum(r["qty"] for r in rows)} pieces')
