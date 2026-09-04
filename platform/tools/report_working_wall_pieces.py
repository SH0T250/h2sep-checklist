import json, re, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
SP=sys.argv[1] if len(sys.argv)>1 else '.'
rows=json.load(open(SP+'/ww_rows.json'))
STAMP='2026-09-03'
OUT=SP+'/H2SEP_Working-Wall-Pieces-On-Site_2026-09-03.xlsx'

# Pieces counted in the field, from the foot of each floor-4 sheet, as written.
COUNTED={
 '430': {'tag':'GR-305','pieces':['L1','L2','L3','L4'],'open':[], 'note':'Prior note on the line: piece GR-305L1-4 missing'},
 '431': {'tag':'GR-304','pieces':['L1','L2','L3','L4','L5','L6'],'open':[], 'note':'Sheet footer: "All boxes in, Not installed"'},
 '433': {'tag':'GR-304','pieces':['R1','R3','R4','R5'],'open':[], 'note':'Prior note on the line: pieces GR-304-4 and GR-304-2 missing'},
 '436': {'tag':'GR-308','pieces':['L1','L2','L3','L4','L5','L7'],'open':['L2'],'note':'L2 box was opened. Footer also reads "Kitchen tile top on floor"'},
 '438': {'tag':'GR-307','pieces':['GR-307-R-1','GR-307-R-2','GR-307-R-4','GR-309-R-5','GR-307-R-6','GR-307-R-7'],'open':[],
         'note':'Piece five is written GR-309-R-5 on the sheet and is left as written'},
}
by_room={r['room']+'|'+r['tag']: r for r in rows}

FONT='Arial'; NAVY=PatternFill('solid',fgColor='1F3A5F')
HDR=Font(name=FONT,bold=True,color='FFFFFF',size=10); BODY=Font(name=FONT,size=10); BOLD=Font(name=FONT,size=10,bold=True)
TITLE=Font(name=FONT,size=15,bold=True,color='1F3A5F'); SUB=Font(name=FONT,size=9,italic=True,color='555555')
KPIL=Font(name=FONT,size=9,bold=True,color='FFFFFF'); KPIV=Font(name=FONT,size=16,bold=True,color='1F3A5F')
WRAP=Alignment(wrap_text=True,vertical='top'); TOP=Alignment(vertical='top'); CEN=Alignment(horizontal='center',vertical='center')
THIN=Side(style='thin',color='D5DBE0'); BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
OPENF=PatternFill('solid',fgColor='FFF3CD'); GREY=PatternFill('solid',fgColor='EEF1F4')

wb=Workbook(); ws=wb.active; ws.title='Working Wall Pieces'
ws['A1']='WORKING WALL PIECES ON SITE, WAITING TO BE INSTALLED'; ws['A1'].font=TITLE
ws['A2']=('Home2 Suites by Hilton, Eagle Pass TX (H2SEP, Triun 24030). Every floor. Working walls whose boxes are in the room and whose pieces were counted in the field. '
          'One row per piece. Counts come from the numbers written at the foot of each room sheet, carried over word for word, and the room and wall come from the live checklist pulled on ' + STAMP + '.')
ws['A2'].font=SUB; ws['A2'].alignment=WRAP; ws.merge_cells('A2:H2'); ws.row_dimensions[2].height=30

pieces=[]
for room,c in COUNTED.items():
    r=by_room.get(room+'|'+c['tag'])
    for p in c['pieces']:
        pieces.append({'floor':r['floor'] if r else '4','room':room,'type':(r['type'] if r else ''),
                       'tag':c['tag'],'piece':p,'open':'Opened' if p in c['open'] else 'Sealed','note':c['note']})
kpis=[('Pieces on site',len(pieces)),('Rooms counted',len(COUNTED)),('Walls counted',len({(k,v['tag']) for k,v in COUNTED.items()})),('Boxes opened',sum(len(v['open']) for v in COUNTED.values()))]
for i,(lab,val) in enumerate(kpis):
    c1=ws.cell(row=4,column=1+i*2,value=lab); c1.font=KPIL; c1.fill=NAVY; c1.alignment=CEN
    ws.merge_cells(start_row=4,start_column=1+i*2,end_row=4,end_column=2+i*2)
    c2=ws.cell(row=5,column=1+i*2,value=val); c2.font=KPIV; c2.alignment=CEN; c2.number_format='#,##0'
    ws.merge_cells(start_row=5,start_column=1+i*2,end_row=5,end_column=2+i*2)
    for col in (1+i*2,2+i*2): ws.cell(row=4,column=col).border=BOX; ws.cell(row=5,column=col).border=BOX
ws.row_dimensions[4].height=18; ws.row_dimensions[5].height=24

H=['Floor','Room','Room Type','Wall Tag','Piece','Box','Pieces In This Room','Note From The Sheet']
hr=7
for c,h in enumerate(H,1):
    cell=ws.cell(row=hr,column=c,value=h); cell.font=HDR; cell.fill=NAVY
    cell.alignment=Alignment(wrap_text=True,vertical='center'); cell.border=BOX
ws.row_dimensions[hr].height=26
i=hr
for p in sorted(pieces, key=lambda x:(int(x['floor']), x['room'], x['piece'])):
    i+=1
    n=len(COUNTED[p['room']]['pieces'])
    vals=[int(p['floor']),p['room'],p['type'],p['tag'],p['piece'],p['open'],n,p['note']]
    for c,v in enumerate(vals,1):
        cell=ws.cell(row=i,column=c,value=v); cell.font=BODY; cell.border=BOX
        cell.alignment=WRAP if c==8 else (Alignment(horizontal='center',vertical='top') if c in (1,6,7) else TOP)
        if c==5: cell.font=BOLD
        if p['open']=='Opened': cell.fill=OPENF
last=i
i+=1
ws.cell(row=i,column=1,value='TOTAL PIECES ON SITE').font=BOLD
ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=4)
cell=ws.cell(row=i,column=5,value=len(pieces)); cell.font=BOLD; cell.alignment=Alignment(horizontal='center')
for c in range(1,9): ws.cell(row=i,column=c).border=BOX
t=Table(displayName='WWPieces', ref='A%d:H%d'%(hr,last))
t.tableStyleInfo=TableStyleInfo(name='TableStyleLight1', showRowStripes=True); ws.add_table(t)

# Second block: walls reported on site but never counted piece by piece.
i+=2
ws.cell(row=i,column=1,value='REPORTED ON SITE BUT NOT COUNTED - nobody has written down the piece numbers in these rooms').font=BOLD
ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=8)
i+=1
H2=['Floor','Room','Room Type','Wall Tag','Piece','Box','Pieces In This Room','What the checklist says']
for c,h in enumerate(H2,1):
    cell=ws.cell(row=i,column=c,value=h); cell.font=HDR; cell.fill=NAVY
    cell.alignment=Alignment(wrap_text=True,vertical='center'); cell.border=BOX
onsite=[]
for r in rows:
    iss=(r['issue'] or '').upper()
    if r['resolved'] or not iss: continue
    if 'NOT INSTALLED:' in iss: continue
    if ('IN BOX' in iss) or ('NEED INSTALL' in iss) or ('PIECES ARE HERE' in iss) or ('ASSEMBLED' in iss):
        onsite.append(r)
for r in sorted(onsite, key=lambda x:(int(x['floor']), x['room'])):
    i+=1
    vals=[int(r['floor']),r['room'],r['type'],r['tag'],'not counted','',None,r['issue']]
    for c,v in enumerate(vals,1):
        cell=ws.cell(row=i,column=c,value=v); cell.font=BODY; cell.border=BOX; cell.fill=GREY
        cell.alignment=WRAP if c==8 else (Alignment(horizontal='center',vertical='top') if c in (1,6,7) else TOP)
for c,w in zip(range(1,9),[7,8,24,10,16,9,11,62]): ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes='A%d'%(hr+1)
i+=2
n=ws.cell(row=i,column=1,value=('How to read this: one row is one box of working wall sitting in that room, waiting on the installer. "Piece" is the number written on the box, carried over exactly as written on the sheet. '
 '"Box" says whether it had been opened. The grey block at the bottom is walls the crew reported on site without writing the piece numbers down, so those rooms still need a count. '
 'Walls already built in are not on this list, and neither are walls reported missing.'))
n.font=SUB; n.alignment=WRAP; ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=8); ws.row_dimensions[i].height=32
wb.save(OUT)
print('pieces on site:',len(pieces),'across',len(COUNTED),'rooms')
print('uncounted walls reported on site:',len(onsite))
for r in sorted(onsite,key=lambda x:(int(x['floor']),x['room'])): print('   F%s %s %s' % (r['floor'],r['room'],r['tag']))
print('saved',OUT)
